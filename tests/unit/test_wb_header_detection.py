from __future__ import annotations

import csv
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.parsers.wb.header_detection import locate_wb_table, load_wb_dataframe
from app.parsers.wb.streaming import iter_wb_normalized_rows


def _write_xlsx(path: Path, *, sheets: dict[str, list[list[object]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets.items():
        sheet = workbook.create_sheet(title)
        for row in rows:
            sheet.append(row)
    workbook.save(path)


def test_locate_header_after_title_rows(tmp_path: Path) -> None:
    path = tmp_path / "wb_title.xlsx"
    _write_xlsx(
        path,
        sheets={
            "Data": [
                ["Еженедельный детализированный отчёт реализации"],
                [],
                [
                    "№",
                    "Дата продажи",
                    "Артикул поставщика",
                    "Код номенклатуры",
                    "Тип операции",
                    "Цена розничная",
                    "К перечислению",
                ],
                ["1", "2026-01-15", "SKU-A", "123", "Продажа", 1000, 800],
            ]
        },
    )
    located = locate_wb_table(path, filename=path.name)
    assert located.header_row_index == 2
    assert located.estimated_data_rows >= 1

    total = sum(len(chunk) for _, chunk in iter_wb_normalized_rows(path, filename=path.name))
    assert total == 1


def test_locate_data_on_non_active_sheet(tmp_path: Path) -> None:
    path = tmp_path / "wb_multisheet.xlsx"
    _write_xlsx(
        path,
        sheets={
            "Summary": [["Сводка отчёта"]],
            "Детализация": [
                ["№", "Дата продажи", "Артикул поставщика", "Код номенклатуры", "Тип операции", "Цена розничная", "К перечислению"],
                ["1", "2026-01-15", "SKU-B", "456", "Продажа", 1500, 1200],
                ["2", "2026-01-16", "SKU-B", "456", "Продажа", 500, 400],
            ],
        },
    )
    located = locate_wb_table(path, filename=path.name)
    assert located.sheet_name == "Детализация"
    total = sum(len(chunk) for _, chunk in iter_wb_normalized_rows(path, filename=path.name))
    assert total == 2


def test_load_wb_dataframe_skips_title_rows() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Заголовок отчёта"])
    sheet.append(["№", "Дата продажи", "Артикул поставщика", "Тип операции", "Цена розничная", "К перечислению"])
    sheet.append(["1", "2026-01-15", "SKU-C", "Продажа", 1000, 900])
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        path = Path(handle.name)
        workbook.save(path)
    try:
        df = load_wb_dataframe(path.name, path.read_bytes())
        assert len(df) == 1
        assert "Дата продажи" in df.columns
    finally:
        path.unlink()


def test_csv_with_preamble_header(tmp_path: Path) -> None:
    path = tmp_path / "wb.csv"
    path.write_text(
        "Отчёт WB\n\n"
        "Дата продажи,Артикул поставщика,Тип операции,Цена розничная,К перечислению\n"
        "2026-01-15,SKU-D,Продажа,1000,800\n",
        encoding="utf-8",
    )
    total = sum(len(chunk) for _, chunk in iter_wb_normalized_rows(path, filename=path.name))
    assert total == 1


def test_real_fixture_still_parses() -> None:
    fixture = Path("tests/Еженедельный детализированный отчет WB.xlsx")
    if not fixture.is_file():
        pytest.skip("fixture missing")
    total = sum(len(chunk) for _, chunk in iter_wb_normalized_rows(fixture, filename=fixture.name))
    assert total == 51
