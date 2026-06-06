"""Detect WB worksheet and header row for heterogeneous Excel exports."""

from __future__ import annotations

import csv
import io
import math
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.parsers.wb.base import normalize_header, resolve_column_map
from app.parsers.wb.mapping import PARSER_VERSION_V1_SIGNATURE, PARSER_VERSION_V2_SIGNATURE

MAX_HEADER_SCAN_ROWS = 50
MIN_SIGNATURE_HITS = 2


@dataclass(frozen=True)
class WbTableLocation:
    sheet_name: str | None
    header_row_index: int
    headers: list[str]
    estimated_data_rows: int = 0


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def row_has_data(values: tuple[object, ...] | list[object]) -> bool:
    return any(not _is_empty(value) for value in values)


def cells_to_headers(cells: tuple[object, ...] | list[object]) -> list[str]:
    return ["" if _is_empty(cell) else str(cell).strip() for cell in cells]


def header_signature_score(headers: list[str]) -> int:
    normalized = {normalize_header(header) for header in headers if header.strip()}
    v1 = len(PARSER_VERSION_V1_SIGNATURE.intersection(normalized))
    v2 = len(PARSER_VERSION_V2_SIGNATURE.intersection(normalized))
    signature_hits = max(v1, v2)
    if signature_hits < MIN_SIGNATURE_HITS:
        ru_markers = (
            "к перечислению",
            "артикул",
            "номенклатур",
            "тип операции",
            "ppvz_for_pay",
            "supplier_oper_name",
            "цена рознич",
        )
        signature_hits = max(
            signature_hits,
            sum(1 for header in normalized if any(marker in header for marker in ru_markers)),
        )
    mapped = sum(1 for value in resolve_column_map(headers).values() if value)
    return signature_hits * 100 + mapped


def locate_wb_table_in_rows(rows: list[tuple[object, ...]]) -> WbTableLocation | None:
    best: tuple[int, int, list[str], int] | None = None
    for index, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS]):
        headers = cells_to_headers(row)
        score = header_signature_score(headers)
        if score < MIN_SIGNATURE_HITS * 100:
            continue
        data_rows = sum(
            1 for candidate in rows[index + 1 : MAX_HEADER_SCAN_ROWS] if row_has_data(candidate)
        )
        total_score = score * 10_000 + data_rows
        if best is None or total_score > best[0]:
            best = (total_score, index, headers, data_rows)
    if best is None:
        return None
    _, header_index, headers, data_rows = best
    return WbTableLocation(
        sheet_name=None,
        header_row_index=header_index,
        headers=headers,
        estimated_data_rows=data_rows,
    )


def _count_data_rows_after_header(sheet, header_row_index: int) -> int:
    count = 0
    for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_index <= header_row_index:
            continue
        if row_has_data(row):
            count += 1
    return count


def locate_wb_table_in_workbook(workbook) -> WbTableLocation | None:
    best: tuple[int, WbTableLocation] | None = None
    for sheet in workbook.worksheets:
        scanned: list[tuple[object, ...]] = []
        for row in sheet.iter_rows(values_only=True):
            scanned.append(row)
            if len(scanned) >= MAX_HEADER_SCAN_ROWS:
                break
        candidate = locate_wb_table_in_rows(scanned)
        if candidate is None:
            continue
        data_rows = _count_data_rows_after_header(sheet, candidate.header_row_index)
        total_score = header_signature_score(candidate.headers) * 10_000 + data_rows
        located = WbTableLocation(
            sheet_name=sheet.title,
            header_row_index=candidate.header_row_index,
            headers=candidate.headers,
            estimated_data_rows=data_rows,
        )
        if best is None or total_score > best[0]:
            best = (total_score, located)
    return best[1] if best else None


def locate_wb_table(path: Path, *, filename: str) -> WbTableLocation:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [tuple(row) for row in csv.reader(handle)]
        located = locate_wb_table_in_rows([tuple(r) for r in rows])
        if located is None:
            if not rows:
                raise ValueError("Report file contains no data rows")
            headers = cells_to_headers(tuple(rows[0]))
            return WbTableLocation(
                sheet_name=None,
                header_row_index=0,
                headers=headers,
                estimated_data_rows=max(0, len(rows) - 1),
            )
        return located

    if suffix not in {".xlsx"}:
        raise ValueError(f"Streaming parse is not supported for {suffix!r}; use .xlsx or .csv")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        located = locate_wb_table_in_workbook(workbook)
        if located is not None:
            return located
        sheet = workbook.active
        try:
            header_row = next(sheet.iter_rows(values_only=True))
        except StopIteration as exc:
            raise ValueError("Report file contains no data rows") from exc
        headers = cells_to_headers(header_row)
        return WbTableLocation(
            sheet_name=sheet.title,
            header_row_index=0,
            headers=headers,
            estimated_data_rows=0,
        )
    finally:
        workbook.close()


def load_wb_dataframe(filename: str, content: bytes) -> pd.DataFrame:
    """Load WB Excel/CSV with worksheet and header-row detection."""
    suffix = Path(filename).suffix.lower()
    buffer = io.BytesIO(content)
    if suffix == ".csv":
        text = buffer.read().decode("utf-8-sig")
        rows = list(csv.reader(text.splitlines()))
        located = locate_wb_table_in_rows([tuple(row) for row in rows])
        if located is None:
            return pd.read_csv(io.StringIO(text))
        data_rows = rows[located.header_row_index + 1 :]
        return pd.DataFrame(data_rows, columns=located.headers)

    if suffix not in {".xlsx", ".xls"}:
        raise ValueError(f"Unsupported file format: {filename}")

    if suffix == ".xlsx":
        buffer.seek(0)
        workbook = load_workbook(buffer, read_only=True, data_only=True)
        try:
            located = locate_wb_table_in_workbook(workbook)
        finally:
            workbook.close()
        if located is None:
            buffer.seek(0)
            return pd.read_excel(buffer)
        buffer.seek(0)
        return pd.read_excel(
            buffer,
            sheet_name=located.sheet_name,
            header=located.header_row_index,
        )

    buffer.seek(0)
    return pd.read_excel(buffer)


def iter_wb_sheet_rows(
    path: Path,
    *,
    filename: str,
) -> Iterator[tuple[WbTableLocation, Iterator[tuple[object, ...]]]]:
    located = locate_wb_table(path, filename=filename)
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for row_index, row in enumerate(reader):
                if row_index <= located.header_row_index:
                    continue
                yield located, iter([tuple(row)])
        return

    with load_workbook(path, read_only=True, data_only=True) as workbook:
        sheet = workbook[located.sheet_name or workbook.active.title]

        def _rows() -> Iterator[tuple[object, ...]]:
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index <= located.header_row_index:
                    continue
                yield row

        yield located, _rows()
