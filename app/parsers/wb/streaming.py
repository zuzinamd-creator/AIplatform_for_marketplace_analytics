"""Memory-bounded WB report row iteration (xlsx/csv on disk)."""

from __future__ import annotations

import csv
import math
from collections.abc import Iterator
from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.parsers.wb.base import (
    NormalizedWbRow,
    WbReportParserStrategy,
    parse_decimal,
    resolve_column_map,
)
from app.parsers.wb.header_detection import locate_wb_table, row_has_data
from app.parsers.wb.mapping import PARSER_VERSION_V2_SIGNATURE
from app.parsers.wb.strategies.realization_v1 import RealizationV1Parser
from app.parsers.wb.strategies.realization_v2 import RealizationV2Parser
from app.parsers.wb.base import normalize_header


def _select_parser_from_headers(headers: list[str]) -> WbReportParserStrategy:
    normalized = {normalize_header(column) for column in headers if column}
    if len(PARSER_VERSION_V2_SIGNATURE.intersection(normalized)) >= 2:
        return RealizationV2Parser()
    return RealizationV1Parser()


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _parse_date_value(value: object) -> date | None:
    if _is_empty(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _row_from_values(
    parser: WbReportParserStrategy,
    *,
    index: int,
    headers: list[str],
    values: tuple[object, ...],
    column_map: dict[str, str | None],
) -> NormalizedWbRow:
    row_map = {
        headers[i]: values[i] if i < len(values) else None for i in range(len(headers))
    }
    raw = {str(k): "" if _is_empty(v) else str(v) for k, v in row_map.items()}
    canonical: dict[str, object] = {}
    for field, column in column_map.items():
        if column is None:
            canonical[field] = None
            continue
        value = row_map.get(column)
        if field == "operation_date":
            canonical[field] = _parse_date_value(value)
        elif field == "sale_date":
            canonical[field] = _parse_date_value(value)
        elif field == "quantity":
            dec = parse_decimal(value)
            canonical[field] = int(dec) if dec is not None else None
        elif field in {"sku", "nm_id", "warehouse_name", "operation_type"}:
            canonical[field] = None if _is_empty(value) else str(value).strip()
        else:
            canonical[field] = parse_decimal(value)

    operation_date = canonical.get("operation_date")
    sku_val = canonical.get("sku")
    nm_val = canonical.get("nm_id")
    return NormalizedWbRow(
        source_row_id=f"row-{index}",
        source_row_index=index,
        operation_date=operation_date if isinstance(operation_date, date) else None,
        sku=str(sku_val) if sku_val else None,
        nm_id=str(nm_val) if nm_val else None,
        canonical=canonical,
        raw=raw,
    )


@contextmanager
def _open_xlsx_workbook(path: Path):
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        yield workbook
    finally:
        workbook.close()


def _yield_normalized_chunks(
    *,
    headers: list[str],
    row_iter: Iterator[tuple[object, ...]],
    chunk_size: int,
) -> Iterator[tuple[WbReportParserStrategy, list[NormalizedWbRow]]]:
    parser = _select_parser_from_headers(headers)
    column_map = resolve_column_map(headers)
    chunk: list[NormalizedWbRow] = []
    index = 0
    for values in row_iter:
        if not row_has_data(values):
            continue
        chunk.append(
            _row_from_values(
                parser,
                index=index,
                headers=headers,
                values=values,
                column_map=column_map,
            )
        )
        index += 1
        if len(chunk) >= chunk_size:
            yield parser, chunk
            chunk = []
    if chunk:
        yield parser, chunk


def iter_wb_normalized_rows(
    path: Path,
    *,
    filename: str,
    chunk_size: int = 1000,
) -> Iterator[tuple[WbReportParserStrategy, list[NormalizedWbRow]]]:
    """
    Yield (parser, chunk) without materializing the full workbook as a DataFrame.

    The workbook is always closed when iteration ends or is interrupted.
    """
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        yield from _iter_csv_chunks(path, chunk_size=chunk_size)
        return
    if suffix not in {".xlsx"}:
        raise ValueError(f"Streaming parse is not supported for {suffix!r}; use .xlsx or .csv")

    located = locate_wb_table(path, filename=filename)
    with _open_xlsx_workbook(path) as workbook:
        sheet = workbook[located.sheet_name or workbook.active.title]
        row_iter = (
            row
            for row_index, row in enumerate(sheet.iter_rows(values_only=True))
            if row_index > located.header_row_index
        )
        yield from _yield_normalized_chunks(
            headers=located.headers,
            row_iter=row_iter,
            chunk_size=chunk_size,
        )


def _iter_csv_chunks(path: Path, *, chunk_size: int) -> Iterator[tuple[WbReportParserStrategy, list[NormalizedWbRow]]]:
    located = locate_wb_table(path, filename=path.name)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        row_iter = (
            tuple(row)
            for row_index, row in enumerate(reader)
            if row_index > located.header_row_index
        )
        yield from _yield_normalized_chunks(
            headers=located.headers,
            row_iter=row_iter,
            chunk_size=chunk_size,
        )
